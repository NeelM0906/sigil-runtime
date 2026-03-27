"""Lightweight document extraction — no heavy ML dependencies."""
from __future__ import annotations

import csv
import io
import logging
from pathlib import Path

log = logging.getLogger(__name__)


def extract_text(file_path: str | Path) -> dict:
    """Extract text from a file using lightweight methods.

    Returns:
        {
            "text": str,
            "format": str,
            "filename": str,
            "byte_size": int,
            "can_send_native": bool,
            "raw_bytes": bytes | None,
        }
    """
    file_path = Path(file_path)
    ext = file_path.suffix.lower()
    byte_size = file_path.stat().st_size

    if ext == ".pdf":
        return _extract_pdf(file_path, byte_size)
    if ext in (".csv", ".tsv"):
        return _extract_csv(file_path, byte_size)
    if ext in (".txt", ".md", ".html"):
        return _extract_text(file_path, byte_size)
    if ext == ".docx":
        return _extract_docx(file_path, byte_size)
    if ext in (".xlsx", ".xls"):
        return _extract_xlsx(file_path, byte_size)
    if ext in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
        return _extract_image(file_path, byte_size)

    # Fallback: try reading as text
    try:
        text = file_path.read_text(encoding="utf-8", errors="replace")
        return {"text": text, "format": ext.lstrip("."), "filename": file_path.name,
                "byte_size": byte_size, "can_send_native": False, "raw_bytes": None}
    except Exception:
        return {"text": f"[Binary file: {file_path.name}, {byte_size} bytes]",
                "format": ext.lstrip("."), "filename": file_path.name,
                "byte_size": byte_size, "can_send_native": False, "raw_bytes": None}


def _extract_pdf(file_path: Path, byte_size: int) -> dict:
    raw_bytes = file_path.read_bytes()
    text = ""
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(raw_bytes))
        pages = []
        for page in reader.pages[:50]:
            pages.append(page.extract_text() or "")
        text = "\n\n".join(pages)
    except Exception as exc:
        log.debug("pypdf extraction failed: %s", exc)

    # OCR fallback for scanned PDFs (no embedded text)
    if not text.strip():
        try:
            from pdf2image import convert_from_bytes
            import pytesseract
            images = convert_from_bytes(raw_bytes, dpi=200, first_page=1, last_page=20)
            ocr_pages = []
            for img in images:
                ocr_text = pytesseract.image_to_string(img, lang="eng")
                if ocr_text.strip():
                    ocr_pages.append(ocr_text.strip())
            if ocr_pages:
                text = "\n\n".join(ocr_pages)
                log.info("OCR extracted %d chars from %d pages of %s", len(text), len(ocr_pages), file_path.name)
        except ImportError:
            log.debug("OCR not available (install pytesseract + pdf2image)")
        except Exception as exc:
            log.debug("OCR failed for %s: %s", file_path.name, exc)

    return {
        "text": text,
        "format": "pdf",
        "filename": file_path.name,
        "byte_size": byte_size,
        "can_send_native": True,
        "raw_bytes": raw_bytes if byte_size <= 30_000_000 else None,
    }


def _extract_csv(file_path: Path, byte_size: int) -> dict:
    text = file_path.read_text(encoding="utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return {"text": "[Empty CSV file]", "format": "csv", "filename": file_path.name,
                "byte_size": byte_size, "can_send_native": False, "raw_bytes": None}
    headers = rows[0]
    md_lines = [
        "| " + " | ".join(h.strip() for h in headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows[1:500]:
        padded = (row + [""] * len(headers))[:len(headers)]
        md_lines.append("| " + " | ".join(c.strip() for c in padded) + " |")
    summary = f"CSV: {file_path.name} ({len(rows)-1} rows, {len(headers)} columns)\n"
    summary += f"Columns: {', '.join(headers)}\n\n"
    summary += "\n".join(md_lines)
    return {"text": summary, "format": "csv", "filename": file_path.name,
            "byte_size": byte_size, "can_send_native": False, "raw_bytes": None}


def _extract_text(file_path: Path, byte_size: int) -> dict:
    text = file_path.read_text(encoding="utf-8", errors="replace")
    return {"text": text, "format": file_path.suffix.lstrip("."), "filename": file_path.name,
            "byte_size": byte_size, "can_send_native": False, "raw_bytes": None}


def _extract_docx(file_path: Path, byte_size: int) -> dict:
    try:
        from docx import Document
        doc = Document(str(file_path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n\n".join(paragraphs)
    except Exception as exc:
        log.warning("DOCX extraction failed: %s", exc)
        text = f"[Could not extract DOCX: {exc}]"
    return {"text": text, "format": "docx", "filename": file_path.name,
            "byte_size": byte_size, "can_send_native": False, "raw_bytes": None}


def _extract_xlsx(file_path: Path, byte_size: int) -> dict:
    ext = file_path.suffix.lower()
    if ext == ".xls":
        return _extract_xls_legacy(file_path, byte_size)
    return _extract_xlsx_openpyxl(file_path, byte_size)


def _extract_xls_legacy(file_path: Path, byte_size: int) -> dict:
    """Extract from old .xls format using xlrd."""
    try:
        import xlrd
        wb = xlrd.open_workbook(str(file_path))
        sheets = []
        for ws in wb.sheets()[:10]:
            rows = []
            for rx in range(min(ws.nrows, 500)):
                rows.append([str(ws.cell_value(rx, cx)) for cx in range(ws.ncols)])
            if rows:
                headers = rows[0]
                md = [
                    f"\n## Sheet: {ws.name}",
                    "| " + " | ".join(headers) + " |",
                    "| " + " | ".join(["---"] * len(headers)) + " |",
                ]
                for r in rows[1:]:
                    padded = (r + [""] * len(headers))[:len(headers)]
                    md.append("| " + " | ".join(padded) + " |")
                sheets.append("\n".join(md))
        text = "\n".join(sheets) if sheets else "[Empty workbook]"
    except ImportError:
        log.debug("xlrd not installed — falling back to openpyxl for .xls")
        return _extract_xlsx_openpyxl(file_path, byte_size)
    except Exception as exc:
        log.warning("XLS extraction failed: %s", exc)
        text = f"[Could not extract XLS: {exc}]"
    return {"text": text, "format": "xls", "filename": file_path.name,
            "byte_size": byte_size, "can_send_native": False, "raw_bytes": None}


def _extract_xlsx_openpyxl(file_path: Path, byte_size: int) -> dict:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        sheets = []
        for ws in wb.worksheets[:10]:
            rows = []
            for row in ws.iter_rows(max_row=500, values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            if rows:
                headers = rows[0]
                md = [
                    f"\n## Sheet: {ws.title}",
                    "| " + " | ".join(headers) + " |",
                    "| " + " | ".join(["---"] * len(headers)) + " |",
                ]
                for r in rows[1:]:
                    padded = (r + [""] * len(headers))[:len(headers)]
                    md.append("| " + " | ".join(padded) + " |")
                sheets.append("\n".join(md))
        text = "\n".join(sheets) if sheets else "[Empty workbook]"
        wb.close()
    except Exception as exc:
        log.warning("XLSX extraction failed: %s", exc)
        text = f"[Could not extract XLSX: {exc}]"
    return {"text": text, "format": "xlsx", "filename": file_path.name,
            "byte_size": byte_size, "can_send_native": False, "raw_bytes": None}


def _extract_image(file_path: Path, byte_size: int) -> dict:
    raw = file_path.read_bytes()
    return {
        "text": f"[Image: {file_path.name}]",
        "format": file_path.suffix.lstrip("."),
        "filename": file_path.name,
        "byte_size": byte_size,
        "can_send_native": True,
        "raw_bytes": raw if byte_size <= 20_000_000 else None,
    }


def chunk_text(text: str, max_chars: int = 3000, overlap: int = 300) -> list[str]:
    """Chunk plain text by paragraph boundaries with overlap."""
    if len(text) <= max_chars:
        return [text]
    chunks: list[str] = []
    start = 0
    while start < len(text):
        end = start + max_chars
        if end < len(text):
            for sep in ['\n\n', '\n', '. ', ' ']:
                pos = text.rfind(sep, start + max_chars // 2, end)
                if pos > start:
                    end = pos + len(sep)
                    break
        chunks.append(text[start:end].strip())
        start = max(end - overlap, start + 1)
    return [c for c in chunks if c]


def chunk_code(text: str, max_chars: int = 3000, overlap: int = 300) -> list[str]:
    """Chunk source code by function/class boundaries.

    Tries to split at function/class definitions so each chunk is a
    self-contained logical unit. Falls back to line-based splitting
    if no boundaries are found.
    """
    import re

    # Patterns that typically start a new logical block in common languages
    boundary_re = re.compile(
        r'^(?:'
        r'(?:public|private|protected|static|abstract|final|async)?\s*'
        r'(?:function|class|interface|trait|def|const|let|var|export|module)\s'
        r'|/\*\*'               # doc-comment blocks
        r'|#{1,3}\s'            # markdown headings (in .md files)
        r'|(?:CREATE|ALTER|DROP)\s'  # SQL DDL
        r')',
        re.MULTILINE | re.IGNORECASE,
    )

    lines = text.split('\n')
    chunks: list[str] = []
    current_lines: list[str] = []
    current_len = 0

    for line in lines:
        is_boundary = bool(boundary_re.match(line.lstrip()))

        # If we hit a boundary and current chunk is big enough, flush
        if is_boundary and current_len > 500:
            chunk = '\n'.join(current_lines).strip()
            if chunk:
                chunks.append(chunk)
            # Keep last few lines as overlap context
            overlap_lines = current_lines[-5:] if len(current_lines) > 5 else []
            current_lines = overlap_lines + [line]
            current_len = sum(len(l) for l in current_lines)
            continue

        current_lines.append(line)
        current_len += len(line) + 1

        # Hard split if chunk gets too large
        if current_len >= max_chars:
            chunk = '\n'.join(current_lines).strip()
            if chunk:
                chunks.append(chunk)
            overlap_lines = current_lines[-5:] if len(current_lines) > 5 else []
            current_lines = overlap_lines
            current_len = sum(len(l) for l in current_lines)

    # Flush remaining
    if current_lines:
        chunk = '\n'.join(current_lines).strip()
        if chunk:
            chunks.append(chunk)

    return chunks if chunks else [text[:max_chars]]



    """Chunk text for embedding storage."""
    paragraphs = text.split("\n\n")
    chunks: list[str] = []
    current = ""
    for para in paragraphs:
        if len(current) + len(para) > max_chars and current:
            chunks.append(current.strip())
            current = current[-overlap:] + "\n\n" + para
        else:
            current = (current + "\n\n" + para) if current else para
    if current.strip():
        chunks.append(current.strip())
    return chunks
